import os
import io
import logging
import zoneinfo
from io import BytesIO
from datetime import datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.conf import settings
from django.core.management import call_command
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from apps.usuarios.views import (
    _estilo_excel, _ajustar_columnas,
    _insertar_encabezado, _nombre_usuario, _formato_fecha,
)
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.date import DateTrigger
from apps.backup.models import Backup, ConfigBackup
from apps.usuarios.models import HistorialCambio
from apps.usuarios.decoradores import admin_requerido
from django_apscheduler.models import DjangoJob

logger = logging.getLogger(__name__)

_POR_PAGINA = 15
_BOGOTA_TZ  = zoneinfo.ZoneInfo('America/Bogota')


@admin_requerido
def lista_backups(request):
    """Lista todos los backups generados — solo ADMIN."""
    backups_qs  = Backup.objects.all()
    paginator   = Paginator(backups_qs, _POR_PAGINA)
    page_obj    = paginator.get_page(request.GET.get('page', 1))

    try:
        job = DjangoJob.objects.get(id='backup_automatico_diario')
        proximo_backup = job.next_run_time
    except DjangoJob.DoesNotExist:
        proximo_backup = None

    breadcrumbs = [
        {'nombre': 'Dashboard',           'url': reverse('core:inicio')},
        {'nombre': 'Copias de seguridad', 'url': None},
    ]
    return render(request, 'backup/lista_backups.html', {
        'backups':            page_obj,
        'page_obj':           page_obj,
        'page_range':         list(paginator.get_elided_page_range(page_obj.number, on_each_side=2, on_ends=1)),
        'paginator_ellipsis': paginator.ELLIPSIS,
        'breadcrumbs':        breadcrumbs,
        'proximo_backup':     proximo_backup,
    })


@admin_requerido
@require_POST
def generar_backup(request):
    """
    Genera un backup manual de la base de datos completa.
    Usa call_command('dumpdata') para evitar problemas de encoding en Windows (DA-005).
    """
    try:
        carpeta_backup = os.path.join(settings.BASE_DIR, 'backups')
        os.makedirs(carpeta_backup, exist_ok=True)

        nombre_archivo = f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        ruta_relativa  = os.path.join('backups', nombre_archivo)
        ruta_absoluta  = os.path.join(settings.BASE_DIR, ruta_relativa)

        output = io.StringIO()
        call_command('dumpdata', indent=2, stdout=output, verbosity=0)

        with open(ruta_absoluta, 'w', encoding='utf-8') as archivo:
            archivo.write(output.getvalue())

        peso_bytes = os.path.getsize(ruta_absoluta)
        if peso_bytes < 1024:
            peso = f'{peso_bytes} B'
        elif peso_bytes < 1024 * 1024:
            peso = f'{peso_bytes / 1024:.1f} KB'
        else:
            peso = f'{peso_bytes / (1024 * 1024):.1f} MB'

        Backup.objects.create(
            nombre_archivo = nombre_archivo,
            tipo           = 'MANUAL',
            peso_archivo   = peso,
            ruta_archivo   = ruta_relativa,
            generado_por   = request.user,
        )

        messages.success(request, f'Backup generado correctamente: {nombre_archivo}')

    except Exception as e:
        logger.error('Error al generar backup: %s', str(e), exc_info=True)
        messages.error(request, 'Error al generar el backup. Revise los logs del servidor.')

    return redirect('backup:lista_backups')


@admin_requerido
def restaurar_backup(request, id):
    """
    Restaura la base de datos desde un archivo de backup.
    Usa call_command('loaddata') — sin subprocess (DA-005).
    Solo accesible para ADMIN — operación crítica.
    """
    backup = get_object_or_404(Backup, id=id)

    if request.method == 'POST':
        try:
            ruta_absoluta = (
                backup.ruta_archivo
                if os.path.isabs(backup.ruta_archivo)
                else os.path.join(settings.BASE_DIR, backup.ruta_archivo)
            )

            if not os.path.exists(ruta_absoluta):
                raise Exception('El archivo de backup no existe en el servidor.')

            call_command('loaddata', ruta_absoluta, verbosity=0)

            HistorialCambio.objects.create(
                modelo        = 'BACKUP',
                objeto_id     = backup.id,
                objeto_nombre = backup.nombre_archivo,
                accion        = 'RESTAURAR',
                observacion   = (
                    f'Base de datos restaurada desde {backup.nombre_archivo} '
                    f'(tipo: {backup.get_tipo_display()}, '
                    f'generado: {backup.fecha_creacion.strftime("%d/%m/%Y %H:%M")})'
                ),
                realizado_por = request.user,
            )

            messages.success(request, f'Base de datos restaurada correctamente desde {backup.nombre_archivo}.')

        except Exception as e:
            logger.error('Error al restaurar backup %s: %s', backup.nombre_archivo, str(e), exc_info=True)
            messages.error(request, 'Error al restaurar el backup. Revise los logs del servidor.')

        return redirect('backup:lista_backups')

    breadcrumbs = [
        {'nombre': 'Dashboard',           'url': reverse('core:inicio')},
        {'nombre': 'Copias de seguridad', 'url': reverse('backup:lista_backups')},
        {'nombre': 'Restaurar backup',    'url': None},
    ]
    return render(request, 'backup/confirmar_restauracion.html', {
        'backup':      backup,
        'breadcrumbs': breadcrumbs,
    })


@admin_requerido
def configuracion_backup(request):
    """Panel de configuración del backup automático — solo ADMIN."""
    from apps.backup.scheduler import get_scheduler, backup_automatico

    config, _ = ConfigBackup.objects.get_or_create(
        pk=1,
        defaults={'hora_diaria': 2, 'minuto_diario': 0, 'activo': True}
    )

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'actualizar_horario':
            try:
                hora   = int(request.POST.get('hora_diaria', 2))
                minuto = int(request.POST.get('minuto_diario', 0))

                if not (0 <= hora <= 23 and 0 <= minuto <= 59):
                    raise ValueError('Hora o minuto fuera de rango.')

                config.hora_diaria    = hora
                config.minuto_diario  = minuto
                config.modificado_por = request.user
                config.save()

                scheduler = get_scheduler()
                if scheduler:
                    scheduler.add_job(
                        backup_automatico,
                        trigger=CronTrigger(hour=hora, minute=minuto),
                        id='backup_automatico_diario',
                        name=f'Backup automático — {hora:02d}:{minuto:02d}',
                        replace_existing=True,
                        misfire_grace_time=3600,
                    )
                    messages.success(request, f'Horario actualizado: backup diario a las {hora:02d}:{minuto:02d}.')
                else:
                    messages.success(request, f'Horario guardado: {hora:02d}:{minuto:02d}. Activo al reiniciar el servidor.')

            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                logger.error('Error actualizando horario backup: %s', str(e), exc_info=True)
                messages.error(request, 'Error al actualizar el horario.')

        elif action == 'programar_puntual':
            try:
                fecha_str = request.POST.get('fecha_puntual', '').strip()
                hora_str  = request.POST.get('hora_puntual', '').strip()

                if not fecha_str or not hora_str:
                    raise ValueError('La fecha y la hora son obligatorias.')

                fecha_hora_naive  = datetime.strptime(f'{fecha_str} {hora_str}', '%Y-%m-%d %H:%M')
                fecha_hora_aware  = fecha_hora_naive.replace(tzinfo=_BOGOTA_TZ)

                if fecha_hora_aware <= timezone.now():
                    raise ValueError('La fecha y hora deben ser futuras.')

                job_id = f'backup_puntual_{fecha_str.replace("-","")}_{hora_str.replace(":","")}'

                scheduler = get_scheduler()
                if not scheduler:
                    raise Exception('El scheduler no está activo. Reinicie el servidor.')

                scheduler.add_job(
                    backup_automatico,
                    trigger=DateTrigger(run_date=fecha_hora_aware),
                    id=job_id,
                    name=f'Backup puntual — {fecha_str} {hora_str}',
                    replace_existing=True,
                    misfire_grace_time=3600,
                )

                messages.success(request, f'Backup puntual programado para el {fecha_str} a las {hora_str}.')

            except ValueError as e:
                messages.error(request, str(e))
            except Exception as e:
                logger.error('Error programando backup puntual: %s', str(e), exc_info=True)
                messages.error(request, str(e) if 'scheduler' in str(e) else 'Error al programar el backup puntual.')

        return redirect('backup:configuracion_backup')

    jobs_puntuales = (
        DjangoJob.objects
        .exclude(id='backup_automatico_diario')
        .filter(next_run_time__isnull=False)
        .order_by('next_run_time')
    )

    breadcrumbs = [
        {'nombre': 'Dashboard',           'url': reverse('core:inicio')},
        {'nombre': 'Copias de seguridad', 'url': reverse('backup:lista_backups')},
        {'nombre': 'Configuración',       'url': None},
    ]
    return render(request, 'backup/configuracion_backup.html', {
        'config':         config,
        'horas':          range(24),
        'minutos':        [0, 15, 30, 45],
        'jobs_puntuales': jobs_puntuales,
        'breadcrumbs':    breadcrumbs,
    })


@admin_requerido
@require_POST
def cancelar_puntual(request):
    """Cancela un backup puntual programado — solo ADMIN."""
    from apps.backup.scheduler import get_scheduler

    job_id = request.POST.get('job_id', '').strip()

    if not job_id.startswith('backup_puntual_'):
        messages.error(request, 'Identificador de job inválido.')
        return redirect('backup:configuracion_backup')

    scheduler = get_scheduler()
    if scheduler:
        try:
            scheduler.remove_job(job_id)
            messages.success(request, 'Backup puntual cancelado correctamente.')
        except Exception:
            messages.error(request, 'No se encontró el backup puntual. Puede que ya haya sido ejecutado.')
    else:
        messages.error(request, 'El scheduler no está activo.')

    return redirect('backup:configuracion_backup')


@admin_requerido
def exportar_backups_excel(request):
    """Genera y descarga el reporte Excel del historial de backups — solo ADMIN."""
    backups = Backup.objects.select_related('generado_por').order_by('-fecha_creacion')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Backups'

    cabecera, dato, borde, color_acento = _estilo_excel(wb)

    encabezados = ['N°', 'Nombre del archivo', 'Tipo', 'Peso', 'Generado por', 'Fecha de generación']
    fila_enc = _insertar_encabezado(
        ws, 'HISTORIAL DE COPIAS DE SEGURIDAD',
        _nombre_usuario(request.user), len(encabezados)
    )
    ws.freeze_panes = f'A{fila_enc + 1}'

    for i, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=fila_enc, column=i, value=enc)
        cell.style = cabecera
    ws.row_dimensions[fila_enc].height = 30

    data_row = fila_enc + 1
    for num, b in enumerate(backups, 1):
        generado_por = _nombre_usuario(b.generado_por) if b.generado_por else 'Automático'
        row_data = [
            num,
            b.nombre_archivo,
            b.get_tipo_display(),
            b.peso_archivo,
            generado_por,
            _formato_fecha(b.fecha_creacion),
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.border = borde
            c.alignment = Alignment(
                vertical='center',
                horizontal='center' if col in (1, 3, 4) else 'left'
            )
        data_row += 1

    total = backups.count()
    ws.cell(row=data_row, column=1, value='Total').font = Font(bold=True, size=10)
    ws.cell(row=data_row, column=1).border = borde
    ws.cell(row=data_row, column=2, value=total).font = Font(bold=True, size=10)
    ws.cell(row=data_row, column=2).border = borde
    ws.cell(row=data_row, column=2).alignment = Alignment(horizontal='center')
    for col in range(3, len(encabezados) + 1):
        ws.cell(row=data_row, column=col).border = borde

    _ajustar_columnas(ws)

    wb.properties.title   = 'Historial de Backups — SOLGASES'
    wb.properties.creator = _nombre_usuario(request.user)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'backups_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    return response
