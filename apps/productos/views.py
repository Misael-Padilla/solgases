from io import BytesIO
from datetime import datetime

from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse
from django.views.decorators.http import require_POST

from apps.productos.models import Producto, HistorialStock
from apps.productos.forms import ProductoForm, StockForm
from apps.usuarios.decoradores import login_requerido, admin_requerido
from apps.usuarios.models import HistorialCambio
from apps.usuarios.views import (
    _estilo_excel, _ajustar_columnas,
    _nombre_usuario, _formato_fecha, _insertar_encabezado,
)

_POR_PAGINA = 15


@login_requerido
def lista_productos(request):
    """Lista todos los productos — accesible para ADMIN y EMP."""
    q = request.GET.get('q', '').strip()
    productos = Producto.objects.all().order_by('codigo')
    if q:
        productos = productos.filter(
            Q(codigo__icontains=q) |
            Q(nombre__icontains=q) |
            Q(categoria__icontains=q)
        )
    paginator = Paginator(productos, _POR_PAGINA)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Productos', 'url': None},
    ]
    return render(request, 'productos/lista_productos.html', {
        'productos':          page_obj,
        'page_obj':           page_obj,
        'page_range':         list(paginator.get_elided_page_range(page_obj.number, on_each_side=2, on_ends=1)),
        'paginator_ellipsis': paginator.ELLIPSIS,
        'q':                  q,
        'breadcrumbs':        breadcrumbs,
    })


@login_requerido
def detalle_producto(request, id):
    """Muestra el detalle de un producto específico."""
    producto  = get_object_or_404(Producto, id=id)
    historial = HistorialStock.objects.filter(
        producto=producto
    ).select_related('realizado_por')[:10]
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Productos', 'url': reverse('productos:lista_productos')},
        {'nombre': producto.nombre, 'url': None},
    ]
    return render(request, 'productos/detalle_producto.html', {
        'producto': producto, 'historial': historial, 'breadcrumbs': breadcrumbs
    })


@login_requerido
def crear_producto(request):
    """Crea un nuevo producto — accesible para ADMIN y EMP."""
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, es_edicion=False)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.creado_por = request.user
            producto.save()
            messages.success(request, 'Producto creado correctamente.')
            return redirect('productos:lista_productos')
    else:
        form = ProductoForm(es_edicion=False)
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Productos', 'url': reverse('productos:lista_productos')},
        {'nombre': 'Crear producto', 'url': None},
    ]
    return render(request, 'productos/form_producto.html', {
        'form': form, 'titulo': 'Crear producto', 'breadcrumbs': breadcrumbs
    })


@login_requerido
def editar_producto(request, id):
    """Edita un producto existente — accesible para ADMIN y EMP."""
    producto = get_object_or_404(Producto, id=id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto, es_edicion=True)
        if form.is_valid():
            p = form.save(commit=False)
            p.modificado_por = request.user
            p.save()
            messages.success(request, 'Producto actualizado correctamente.')
            return redirect('productos:detalle_producto', id=producto.id)
    else:
        form = ProductoForm(instance=producto, es_edicion=True)
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Productos', 'url': reverse('productos:lista_productos')},
        {'nombre': producto.nombre, 'url': reverse('productos:detalle_producto', args=[producto.id])},
        {'nombre': 'Editar', 'url': None},
    ]
    return render(request, 'productos/form_producto.html', {
        'form': form, 'titulo': 'Editar producto', 'breadcrumbs': breadcrumbs
    })


@admin_requerido
@require_POST
def cambiar_estado_producto(request, id):
    """Activa o desactiva un producto — solo ADMIN. Requiere observación."""
    producto = get_object_or_404(Producto, id=id)
    observacion = request.POST.get('observacion', '').strip()
    if not observacion:
        messages.error(request, 'La observación es obligatoria para cambiar el estado.')
        return redirect('productos:lista_productos')
    if producto.estado == 'ACTIVO':
        producto.estado = 'INACTIVO'
        accion = 'DESACTIVAR'
        messages.success(request, f'Producto {producto.nombre} desactivado.')
    else:
        producto.estado = 'ACTIVO'
        accion = 'ACTIVAR'
        messages.success(request, f'Producto {producto.nombre} activado.')
    producto.modificado_por = request.user
    producto.save()
    HistorialCambio.objects.create(
        modelo='PRODUCTO',
        objeto_id=producto.id,
        objeto_nombre=producto.nombre,
        accion=accion,
        observacion=observacion,
        realizado_por=request.user,
    )
    return redirect('productos:lista_productos')


@login_requerido
def exportar_productos_excel(request):
    """Genera y descarga el reporte Excel de productos."""
    import openpyxl
    from openpyxl.styles import Alignment

    productos = Producto.objects.select_related(
        'creado_por', 'modificado_por'
    ).order_by('codigo')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'

    cabecera, dato, borde, _ = _estilo_excel(wb)

    encabezados = [
        'Código', 'Nombre', 'Categoría', 'Género', 'Talla',
        'Precio compra', 'Precio venta',
        'Stock', 'Stock mínimo', 'Stock máximo', 'Nivel stock',
        'Estado', 'Fecha registro',
        'Creado por', 'Modificado por', 'Última modificación',
    ]
    fila_enc = _insertar_encabezado(
        ws, 'REPORTE DE PRODUCTOS', _nombre_usuario(request.user), len(encabezados)
    )
    ws.freeze_panes = f'A{fila_enc + 1}'

    for i, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=fila_enc, column=i, value=enc)
        cell.style = cabecera
    ws.row_dimensions[fila_enc].height = 30

    data_row = fila_enc + 1
    for p in productos:
        row_data = [
            p.codigo,
            p.nombre,
            p.get_categoria_display(),
            p.genero,
            p.talla or '—',
            float(p.precio_compra),
            float(p.precio_venta),
            p.stock,
            p.stock_minimo,
            p.stock_maximo,
            p.nivel_stock,
            p.estado,
            _formato_fecha(p.fecha_creacion),
            _nombre_usuario(p.creado_por),
            _nombre_usuario(p.modificado_por),
            _formato_fecha(p.modificado_en),
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.border = borde
            c.alignment = Alignment(vertical='center')
        data_row += 1

    _ajustar_columnas(ws)
    wb.properties.title   = 'Reporte de Productos — SOLGASES'
    wb.properties.creator = _nombre_usuario(request.user)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    return response


@admin_requerido
def modificar_stock(request, id):
    """Modifica el stock manualmente — solo ADMIN. Requiere motivo obligatorio."""
    producto = get_object_or_404(Producto, id=id)
    if request.method == 'POST':
        form = StockForm(request.POST)
        if form.is_valid():
            stock_anterior      = producto.stock
            producto.stock      = form.cleaned_data['stock']
            producto.modificado_por = request.user
            producto.save()
            HistorialStock.objects.create(
                producto=producto,
                tipo='MANUAL',
                stock_anterior=stock_anterior,
                stock_nuevo=producto.stock,
                motivo=form.cleaned_data['motivo'],
                realizado_por=request.user,
            )
            messages.success(request, f'Stock de {producto.nombre} actualizado a {producto.stock}.')
            return redirect('productos:detalle_producto', id=producto.id)
    else:
        form = StockForm(initial={'stock': producto.stock})
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Productos', 'url': reverse('productos:lista_productos')},
        {'nombre': producto.nombre, 'url': reverse('productos:detalle_producto', args=[producto.id])},
        {'nombre': 'Modificar stock', 'url': None},
    ]
    return render(request, 'productos/form_stock.html', {
        'form': form, 'producto': producto, 'breadcrumbs': breadcrumbs
    })
