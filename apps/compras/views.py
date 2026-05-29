from io import BytesIO
from datetime import datetime

from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse
from django.views.decorators.http import require_POST

from apps.compras.models import FacturaCompra
from apps.compras.forms import FacturaCompraForm, DetalleCompraFormSet
from apps.productos.models import Producto, HistorialStock
from apps.insumos.models import Insumo, HistorialStockInsumo
from apps.usuarios.decoradores import login_requerido, admin_requerido
from apps.usuarios.models import HistorialCambio
from apps.usuarios.views import (
    _estilo_excel, _ajustar_columnas,
    _nombre_usuario, _formato_fecha, _insertar_encabezado,
)

_POR_PAGINA = 15


@login_requerido
def lista_compras(request):
    """Lista todas las facturas de compra — accesible para ADMIN y EMP."""
    q = request.GET.get('q', '').strip()
    compras = FacturaCompra.objects.all().order_by('-fecha_registro')
    if q:
        compras = compras.filter(
            Q(numero_factura__icontains=q) |
            Q(proveedor__razon_social__icontains=q) |
            Q(proveedor__nombres__icontains=q)
        )
    paginator = Paginator(compras, _POR_PAGINA)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Compras', 'url': None},
    ]
    return render(request, 'compras/lista_compras.html', {
        'compras':            page_obj,
        'page_obj':           page_obj,
        'page_range':         list(paginator.get_elided_page_range(page_obj.number, on_each_side=2, on_ends=1)),
        'paginator_ellipsis': paginator.ELLIPSIS,
        'q':                  q,
        'breadcrumbs':        breadcrumbs,
    })


@login_requerido
def detalle_compra(request, id):
    """Muestra el detalle de una factura de compra específica."""
    compra = get_object_or_404(FacturaCompra, id=id)
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Compras', 'url': reverse('compras:lista_compras')},
        {'nombre': compra.numero_factura, 'url': None},
    ]
    return render(request, 'compras/detalle_compra.html', {
        'compra': compra, 'breadcrumbs': breadcrumbs
    })


@login_requerido
def crear_compra(request):
    """Registra una nueva factura de compra con sus detalles."""
    if request.method == 'POST':
        form    = FacturaCompraForm(request.POST)
        formset = DetalleCompraFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    compra = form.save(commit=False)
                    compra.registrado_por = request.user
                    compra.save()
                    formset.instance = compra
                    formset.save()
                    messages.success(request, f'Factura {compra.numero_factura} registrada correctamente.')
                    return redirect('compras:detalle_compra', id=compra.id)
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form    = FacturaCompraForm()
        formset = DetalleCompraFormSet()
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Compras', 'url': reverse('compras:lista_compras')},
        {'nombre': 'Registrar factura', 'url': None},
    ]
    return render(request, 'compras/form_compra.html', {
        'form':    form,
        'formset': formset,
        'titulo':  'Registrar factura de compra',
        'breadcrumbs': breadcrumbs,
    })


@admin_requerido
@require_POST
def cambiar_estado_compra(request, id):
    """
    Activa o desactiva una factura — solo ADMIN. Requiere observación.
    Al desactivar: revierte el stock de todos los ítems (Pendiente Fase 5).
    Al activar: re-incrementa el stock de todos los ítems.
    """
    compra = get_object_or_404(FacturaCompra, id=id)
    observacion = request.POST.get('observacion', '').strip()
    if not observacion:
        messages.error(request, 'La observación es obligatoria para cambiar el estado.')
        return redirect('compras:lista_compras')

    with transaction.atomic():
        if compra.estado == 'ACTIVO':
            compra.estado = 'INACTIVO'
            accion        = 'DESACTIVAR'
            delta         = -1
            msg_stock     = 'revertido'
        else:
            compra.estado = 'ACTIVO'
            accion        = 'ACTIVAR'
            delta         = 1
            msg_stock     = 'restaurado'

        # Revertir / restaurar stock de cada ítem de la factura
        tipo_mov = 'anulación' if delta == -1 else 'reactivación'
        for detalle in compra.detalles.all():
            motivo_stock = (
                f'{"Reversión" if delta == -1 else "Restauración"} por '
                f'{tipo_mov} de compra {compra.numero_factura}: {observacion}'
            )
            if detalle.tipo_item == 'PRODUCTO':
                try:
                    producto  = Producto.objects.get(codigo=detalle.codigo_item)
                    stock_ant = producto.stock
                    producto.stock = max(0, producto.stock + delta * detalle.cantidad)
                    producto.save()
                    HistorialStock.objects.create(
                        producto=producto, tipo='MANUAL',
                        stock_anterior=stock_ant,
                        stock_nuevo=producto.stock,
                        motivo=motivo_stock,
                        realizado_por=request.user,
                    )
                except Producto.DoesNotExist:
                    pass
            elif detalle.tipo_item == 'INSUMO':
                try:
                    insumo    = Insumo.objects.get(codigo=detalle.codigo_item)
                    stock_ant = insumo.stock
                    insumo.stock = max(0, insumo.stock + delta * detalle.cantidad)
                    insumo.save()
                    HistorialStockInsumo.objects.create(
                        insumo=insumo, tipo='MANUAL',
                        stock_anterior=stock_ant,
                        stock_nuevo=insumo.stock,
                        motivo=motivo_stock,
                        realizado_por=request.user,
                    )
                except Insumo.DoesNotExist:
                    pass

        compra.save()
        HistorialCambio.objects.create(
            modelo='COMPRA',
            objeto_id=compra.id,
            objeto_nombre=compra.numero_factura,
            accion=accion,
            observacion=observacion,
            realizado_por=request.user,
        )

    verb = 'desactivada' if accion == 'DESACTIVAR' else 'activada'
    messages.success(request, f'Factura {compra.numero_factura} {verb}. Stock {msg_stock}.')
    return redirect('compras:lista_compras')


@login_requerido
def exportar_compras_excel(request):
    """Genera y descarga el reporte Excel de facturas de compra."""
    import openpyxl
    from openpyxl.styles import Alignment

    compras = FacturaCompra.objects.select_related(
        'proveedor', 'registrado_por'
    ).order_by('-fecha_registro')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Compras'

    cabecera, dato, borde, _ = _estilo_excel(wb)

    encabezados = [
        'N° Factura', 'Proveedor', 'Registrado por',
        'Fecha factura', 'Fecha registro',
        'Subtotal', 'IVA (%)', 'IVA ($)', 'Total',
        'Estado', 'Observaciones',
    ]
    fila_enc = _insertar_encabezado(
        ws, 'REPORTE DE COMPRAS', _nombre_usuario(request.user), len(encabezados)
    )
    ws.freeze_panes = f'A{fila_enc + 1}'

    for i, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=fila_enc, column=i, value=enc)
        cell.style = cabecera
    ws.row_dimensions[fila_enc].height = 30

    data_row = fila_enc + 1
    for c in compras:
        row_data = [
            c.numero_factura,
            str(c.proveedor),
            _nombre_usuario(c.registrado_por),
            _formato_fecha(c.fecha_factura),
            _formato_fecha(c.fecha_registro),
            float(c.subtotal),
            float(c.iva_porcentaje),
            float(c.iva),
            float(c.total),
            c.estado,
            c.observaciones or '—',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.border = borde
            cell.alignment = Alignment(vertical='center')
        data_row += 1

    _ajustar_columnas(ws)
    wb.properties.title   = 'Reporte de Compras — SOLGASES'
    wb.properties.creator = _nombre_usuario(request.user)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'compras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    return response
