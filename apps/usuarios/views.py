from io import BytesIO
from datetime import datetime

from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

from apps.usuarios.models import Usuario, Cliente, Proveedor, HistorialCambio
from apps.usuarios.decoradores import login_requerido, admin_requerido


# =====================================================
# VISTAS DE USUARIOS (ADMIN / EMP)
# =====================================================

@login_requerido
def lista_usuarios(request):
    """Lista todos los usuarios del sistema — accesible para ADMIN y EMP."""
    q = request.GET.get('q', '').strip()
    usuarios = Usuario.objects.all().order_by('nombres')
    if q:
        usuarios = usuarios.filter(
            Q(identificacion__icontains=q) |
            Q(nombres__icontains=q) |
            Q(apellidos__icontains=q) |
            Q(correo_electronico__icontains=q)
        )
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Usuarios', 'url': None},
    ]
    return render(request, 'usuarios/lista_usuarios.html', {'usuarios': usuarios, 'q': q, 'breadcrumbs': breadcrumbs})


@login_requerido
def detalle_usuario(request, id):
    """Muestra el detalle de un usuario específico."""
    usuario = get_object_or_404(Usuario, id=id)
    historial = HistorialCambio.objects.filter(
        modelo='USUARIO', objeto_id=usuario.id
    ).select_related('realizado_por')[:10]
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Usuarios', 'url': reverse('usuarios:lista_usuarios')},
        {'nombre': f'{usuario.nombres} {usuario.apellidos}', 'url': None},
    ]
    return render(request, 'usuarios/detalle_usuario.html', {
        'usuario': usuario, 'historial': historial, 'breadcrumbs': breadcrumbs
    })


@admin_requerido
def crear_usuario(request):
    """Crea un nuevo usuario — solo ADMIN."""
    from apps.usuarios.forms import UsuarioForm
    if request.method == 'POST':
        form = UsuarioForm(request.POST, request.FILES)
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.set_password(form.cleaned_data['password'])
            usuario.creado_por = request.user
            usuario.save()
            messages.success(request, 'Usuario creado correctamente.')
            return redirect('usuarios:lista_usuarios')
    else:
        form = UsuarioForm()
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Usuarios', 'url': reverse('usuarios:lista_usuarios')},
        {'nombre': 'Crear usuario', 'url': None},
    ]
    return render(request, 'usuarios/form_usuario.html', {'form': form, 'titulo': 'Crear usuario', 'breadcrumbs': breadcrumbs})


@admin_requerido
def editar_usuario(request, id):
    """Edita un usuario existente — solo ADMIN."""
    from apps.usuarios.forms import UsuarioEditarForm
    usuario = get_object_or_404(Usuario, id=id)
    if request.method == 'POST':
        form = UsuarioEditarForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            u = form.save(commit=False)
            u.modificado_por = request.user
            u.save()
            observacion_cambio = request.POST.get('observacion_cambio', '').strip()
            if observacion_cambio:
                HistorialCambio.objects.create(
                    modelo='USUARIO',
                    objeto_id=u.id,
                    objeto_nombre=f'{u.nombres} {u.apellidos}',
                    accion='EDITAR',
                    observacion=observacion_cambio,
                    realizado_por=request.user,
                )
            messages.success(request, 'Usuario actualizado correctamente.')
            return redirect('usuarios:detalle_usuario', id=usuario.id)
    else:
        form = UsuarioEditarForm(instance=usuario)
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Usuarios', 'url': reverse('usuarios:lista_usuarios')},
        {'nombre': f'{usuario.nombres} {usuario.apellidos}', 'url': reverse('usuarios:detalle_usuario', args=[usuario.id])},
        {'nombre': 'Editar', 'url': None},
    ]
    return render(request, 'usuarios/form_usuario.html', {
        'form': form, 'titulo': 'Editar usuario', 'es_edicion': True, 'breadcrumbs': breadcrumbs
    })


@admin_requerido
@require_POST
def cambiar_estado_usuario(request, id):
    """Activa o desactiva un usuario — solo ADMIN. Requiere observación."""
    usuario = get_object_or_404(Usuario, id=id)
    observacion = request.POST.get('observacion', '').strip()
    if not observacion:
        messages.error(request, 'La observación es obligatoria para cambiar el estado.')
        return redirect('usuarios:lista_usuarios')
    if usuario.estado == 'ACTIVO':
        usuario.estado = 'INACTIVO'
        accion = 'DESACTIVAR'
        messages.success(request, f'Usuario {usuario.nombres} desactivado.')
    else:
        usuario.estado = 'ACTIVO'
        accion = 'ACTIVAR'
        messages.success(request, f'Usuario {usuario.nombres} activado.')
    usuario.save()
    HistorialCambio.objects.create(
        modelo='USUARIO',
        objeto_id=usuario.id,
        objeto_nombre=f'{usuario.nombres} {usuario.apellidos}',
        accion=accion,
        observacion=observacion,
        realizado_por=request.user,
    )
    return redirect('usuarios:lista_usuarios')


# =====================================================
# VISTAS DE CLIENTES
# =====================================================

@login_requerido
def lista_clientes(request):
    """Lista todos los clientes — accesible para ADMIN y EMP."""
    q = request.GET.get('q', '').strip()
    clientes = Cliente.objects.all().order_by('identificacion')
    if q:
        clientes = clientes.filter(
            Q(identificacion__icontains=q) |
            Q(nombres__icontains=q) |
            Q(apellidos__icontains=q) |
            Q(razon_social__icontains=q) |
            Q(ciudad__icontains=q) |
            Q(telefono__icontains=q)
        )
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Clientes', 'url': None},
    ]
    return render(request, 'usuarios/lista_clientes.html', {'clientes': clientes, 'q': q, 'breadcrumbs': breadcrumbs})


@login_requerido
def detalle_cliente(request, id):
    """Muestra el detalle de un cliente específico."""
    cliente = get_object_or_404(Cliente, id=id)
    nombre_cliente = cliente.razon_social if cliente.tipo_identificacion == 'NIT' else f'{cliente.nombres} {cliente.apellidos}'
    historial = HistorialCambio.objects.filter(
        modelo='CLIENTE', objeto_id=cliente.id
    ).select_related('realizado_por')[:10]
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Clientes', 'url': reverse('usuarios:lista_clientes')},
        {'nombre': nombre_cliente, 'url': None},
    ]
    return render(request, 'usuarios/detalle_cliente.html', {
        'cliente': cliente, 'historial': historial, 'breadcrumbs': breadcrumbs
    })


@login_requerido
def crear_cliente(request):
    """Crea un nuevo cliente — accesible para ADMIN y EMP."""
    from apps.usuarios.forms import ClienteForm
    if request.method == 'POST':
        form = ClienteForm(request.POST, request.FILES)
        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.creado_por = request.user
            cliente.save()
            messages.success(request, 'Cliente creado correctamente.')
            return redirect('usuarios:lista_clientes')
    else:
        form = ClienteForm()
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Clientes', 'url': reverse('usuarios:lista_clientes')},
        {'nombre': 'Crear cliente', 'url': None},
    ]
    return render(request, 'usuarios/form_cliente.html', {'form': form, 'titulo': 'Crear cliente', 'breadcrumbs': breadcrumbs})


@login_requerido
def editar_cliente(request, id):
    """Edita un cliente existente — accesible para ADMIN y EMP."""
    from apps.usuarios.forms import ClienteForm
    cliente = get_object_or_404(Cliente, id=id)
    if request.method == 'POST':
        form = ClienteForm(request.POST, request.FILES, instance=cliente)
        if form.is_valid():
            c = form.save(commit=False)
            c.modificado_por = request.user
            c.save()
            observacion_cambio = request.POST.get('observacion_cambio', '').strip()
            nombre_obj = c.razon_social if c.tipo_identificacion == 'NIT' else f'{c.nombres} {c.apellidos}'
            if observacion_cambio:
                HistorialCambio.objects.create(
                    modelo='CLIENTE',
                    objeto_id=c.id,
                    objeto_nombre=nombre_obj,
                    accion='EDITAR',
                    observacion=observacion_cambio,
                    realizado_por=request.user,
                )
            messages.success(request, 'Cliente actualizado correctamente.')
            return redirect('usuarios:detalle_cliente', id=cliente.id)
    else:
        form = ClienteForm(instance=cliente)
    nombre_cliente = cliente.razon_social if cliente.tipo_identificacion == 'NIT' else f'{cliente.nombres} {cliente.apellidos}'
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Clientes', 'url': reverse('usuarios:lista_clientes')},
        {'nombre': nombre_cliente, 'url': reverse('usuarios:detalle_cliente', args=[cliente.id])},
        {'nombre': 'Editar', 'url': None},
    ]
    return render(request, 'usuarios/form_cliente.html', {
        'form': form, 'titulo': 'Editar cliente', 'es_edicion': True, 'breadcrumbs': breadcrumbs
    })


@admin_requerido
@require_POST
def cambiar_estado_cliente(request, id):
    """Activa o desactiva un cliente — solo ADMIN. Requiere observación."""
    cliente = get_object_or_404(Cliente, id=id)
    observacion = request.POST.get('observacion', '').strip()
    if not observacion:
        messages.error(request, 'La observación es obligatoria para cambiar el estado.')
        return redirect('usuarios:lista_clientes')
    nombre_obj = cliente.razon_social if cliente.tipo_identificacion == 'NIT' else f'{cliente.nombres} {cliente.apellidos}'
    if cliente.estado == 'ACTIVO':
        cliente.estado = 'INACTIVO'
        accion = 'DESACTIVAR'
        messages.success(request, 'Cliente desactivado.')
    else:
        cliente.estado = 'ACTIVO'
        accion = 'ACTIVAR'
        messages.success(request, 'Cliente activado.')
    cliente.save()
    HistorialCambio.objects.create(
        modelo='CLIENTE',
        objeto_id=cliente.id,
        objeto_nombre=nombre_obj,
        accion=accion,
        observacion=observacion,
        realizado_por=request.user,
    )
    return redirect('usuarios:lista_clientes')


# =====================================================
# VISTAS DE PROVEEDORES
# =====================================================

@login_requerido
def lista_proveedores(request):
    """Lista todos los proveedores — accesible para ADMIN y EMP."""
    q = request.GET.get('q', '').strip()
    proveedores = Proveedor.objects.all().order_by('identificacion')
    if q:
        proveedores = proveedores.filter(
            Q(identificacion__icontains=q) |
            Q(nombres__icontains=q) |
            Q(apellidos__icontains=q) |
            Q(razon_social__icontains=q) |
            Q(ciudad__icontains=q) |
            Q(telefono__icontains=q)
        )
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Proveedores', 'url': None},
    ]
    return render(request, 'usuarios/lista_proveedores.html', {'proveedores': proveedores, 'q': q, 'breadcrumbs': breadcrumbs})


@login_requerido
def detalle_proveedor(request, id):
    """Muestra el detalle de un proveedor específico."""
    proveedor = get_object_or_404(Proveedor, id=id)
    nombre_proveedor = proveedor.razon_social if proveedor.tipo_identificacion == 'NIT' else f'{proveedor.nombres} {proveedor.apellidos}'
    historial = HistorialCambio.objects.filter(
        modelo='PROVEEDOR', objeto_id=proveedor.id
    ).select_related('realizado_por')[:10]
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Proveedores', 'url': reverse('usuarios:lista_proveedores')},
        {'nombre': nombre_proveedor, 'url': None},
    ]
    return render(request, 'usuarios/detalle_proveedor.html', {
        'proveedor': proveedor, 'historial': historial, 'breadcrumbs': breadcrumbs
    })


@login_requerido
def crear_proveedor(request):
    """Crea un nuevo proveedor — accesible para ADMIN y EMP."""
    from apps.usuarios.forms import ProveedorForm
    if request.method == 'POST':
        form = ProveedorForm(request.POST, request.FILES)
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.creado_por = request.user
            proveedor.save()
            messages.success(request, 'Proveedor creado correctamente.')
            return redirect('usuarios:lista_proveedores')
    else:
        form = ProveedorForm()
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Proveedores', 'url': reverse('usuarios:lista_proveedores')},
        {'nombre': 'Crear proveedor', 'url': None},
    ]
    return render(request, 'usuarios/form_proveedor.html', {'form': form, 'titulo': 'Crear proveedor', 'breadcrumbs': breadcrumbs})


@login_requerido
def editar_proveedor(request, id):
    """Edita un proveedor existente — accesible para ADMIN y EMP."""
    from apps.usuarios.forms import ProveedorForm
    proveedor = get_object_or_404(Proveedor, id=id)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, request.FILES, instance=proveedor)
        if form.is_valid():
            p = form.save(commit=False)
            p.modificado_por = request.user
            p.save()
            observacion_cambio = request.POST.get('observacion_cambio', '').strip()
            nombre_obj = p.razon_social if p.tipo_identificacion == 'NIT' else f'{p.nombres} {p.apellidos}'
            if observacion_cambio:
                HistorialCambio.objects.create(
                    modelo='PROVEEDOR',
                    objeto_id=p.id,
                    objeto_nombre=nombre_obj,
                    accion='EDITAR',
                    observacion=observacion_cambio,
                    realizado_por=request.user,
                )
            messages.success(request, 'Proveedor actualizado correctamente.')
            return redirect('usuarios:detalle_proveedor', id=proveedor.id)
    else:
        form = ProveedorForm(instance=proveedor)
    nombre_proveedor = proveedor.razon_social if proveedor.tipo_identificacion == 'NIT' else f'{proveedor.nombres} {proveedor.apellidos}'
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Proveedores', 'url': reverse('usuarios:lista_proveedores')},
        {'nombre': nombre_proveedor, 'url': reverse('usuarios:detalle_proveedor', args=[proveedor.id])},
        {'nombre': 'Editar', 'url': None},
    ]
    return render(request, 'usuarios/form_proveedor.html', {
        'form': form, 'titulo': 'Editar proveedor', 'es_edicion': True, 'breadcrumbs': breadcrumbs
    })


@admin_requerido
@require_POST
def cambiar_estado_proveedor(request, id):
    """Activa o desactiva un proveedor — solo ADMIN. Requiere observación."""
    proveedor = get_object_or_404(Proveedor, id=id)
    observacion = request.POST.get('observacion', '').strip()
    if not observacion:
        messages.error(request, 'La observación es obligatoria para cambiar el estado.')
        return redirect('usuarios:lista_proveedores')
    nombre_obj = proveedor.razon_social if proveedor.tipo_identificacion == 'NIT' else f'{proveedor.nombres} {proveedor.apellidos}'
    if proveedor.estado == 'ACTIVO':
        proveedor.estado = 'INACTIVO'
        accion = 'DESACTIVAR'
        messages.success(request, 'Proveedor desactivado.')
    else:
        proveedor.estado = 'ACTIVO'
        accion = 'ACTIVAR'
        messages.success(request, 'Proveedor activado.')
    proveedor.save()
    HistorialCambio.objects.create(
        modelo='PROVEEDOR',
        objeto_id=proveedor.id,
        objeto_nombre=nombre_obj,
        accion=accion,
        observacion=observacion,
        realizado_por=request.user,
    )
    return redirect('usuarios:lista_proveedores')


# =====================================================
# EXPORTACIÓN EXCEL — AUDITORÍA
# =====================================================

def _estilo_excel(wb):
    """Devuelve los estilos reutilizables para los reportes Excel."""
    color_cabecera = '1A1A1A'
    color_acento   = 'CC0000'
    borde = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    cabecera = openpyxl.styles.NamedStyle(name='cabecera')
    cabecera.font      = Font(bold=True, color='FFFFFF', size=11)
    cabecera.fill      = PatternFill('solid', fgColor=color_cabecera)
    cabecera.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cabecera.border    = borde
    dato = openpyxl.styles.NamedStyle(name='dato')
    dato.font      = Font(size=10)
    dato.alignment = Alignment(vertical='center')
    dato.border    = borde
    return cabecera, dato, borde, color_acento


def _ajustar_columnas(ws):
    """Ajusta el ancho de cada columna al contenido más largo."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _nombre_usuario(usuario):
    if usuario:
        return f'{usuario.nombres} {usuario.apellidos}'
    return '—'


def _formato_fecha(dt):
    if dt:
        from zoneinfo import ZoneInfo
        from django.utils import timezone
        if timezone.is_aware(dt):
            dt = dt.astimezone(ZoneInfo('America/Bogota'))
        return dt.strftime('%d/%m/%Y %H:%M')
    return '—'


def _insertar_encabezado(ws, titulo, exportado_por, num_columnas):
    """Inserta logo, título del reporte y metadatos en las filas 1-4."""
    from zoneinfo import ZoneInfo
    ultima_col = get_column_letter(num_columnas)

    # Fila 1: Logo
    ws.row_dimensions[1].height = 55
    logo_path = settings.BASE_DIR / 'static' / 'img' / 'logo_solgases.png'
    if logo_path.exists():
        img = XlImage(str(logo_path))
        img.width = 180
        img.height = 65
        ws.add_image(img, 'A1')

    # Fila 2: Título
    ws.merge_cells(f'A2:{ultima_col}2')
    ws['A2'] = titulo
    ws['A2'].font = Font(bold=True, size=14, color='1A1A1A')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # Fila 3: Metadatos (fecha y usuario que exportó)
    ws.merge_cells(f'A3:{ultima_col}3')
    ahora = datetime.now(ZoneInfo('America/Bogota')).strftime('%d/%m/%Y %H:%M')
    ws['A3'] = f'Generado: {ahora}  |  Exportado por: {exportado_por}'
    ws['A3'].font = Font(size=9, color='666666', italic=True)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 16

    # Fila 4: Separador visual
    ws['A4'] = ''
    ws.row_dimensions[4].height = 6

    return 5  # Fila donde van los encabezados de columna


@login_requerido
def exportar_usuarios_excel(request):
    """Genera y descarga el reporte Excel de usuarios con auditoría."""
    usuarios = Usuario.objects.select_related('creado_por', 'modificado_por').order_by('nombres')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Usuarios'

    cabecera, dato, borde, _ = _estilo_excel(wb)

    encabezados = [
        'Identificación', 'Nombre', 'Correo', 'Rol', 'Estado',
        'Creado por', 'Fecha creación', 'Modificado por', 'Última modificación'
    ]
    fila_enc = _insertar_encabezado(ws, 'REPORTE DE USUARIOS', _nombre_usuario(request.user), len(encabezados))
    ws.freeze_panes = f'A{fila_enc + 1}'

    for i, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=fila_enc, column=i, value=enc)
        cell.style = cabecera
    ws.row_dimensions[fila_enc].height = 30

    data_row = fila_enc + 1
    for u in usuarios:
        row_data = [
            f'{u.tipo_identificacion} {u.identificacion}',
            f'{u.nombres} {u.apellidos}',
            u.correo_electronico,
            u.rol,
            u.estado,
            _nombre_usuario(u.creado_por),
            _formato_fecha(u.fecha_creacion),
            _nombre_usuario(u.modificado_por),
            _formato_fecha(u.modificado_en),
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.border = borde
            c.alignment = Alignment(vertical='center')
        data_row += 1

    _ajustar_columnas(ws)

    wb.properties.title   = 'Reporte de Usuarios — SOLGASES'
    wb.properties.creator = _nombre_usuario(request.user)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    return response


@login_requerido
def exportar_clientes_excel(request):
    """Genera y descarga el reporte Excel de clientes con auditoría."""
    clientes = Cliente.objects.select_related('creado_por', 'modificado_por').order_by('identificacion')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clientes'

    cabecera, dato, borde, _ = _estilo_excel(wb)

    encabezados = [
        'Identificación', 'Nombre / Razón social', 'Teléfono', 'Ciudad',
        'Estado', 'Creado por', 'Fecha creación', 'Modificado por', 'Última modificación'
    ]
    fila_enc = _insertar_encabezado(ws, 'REPORTE DE CLIENTES', _nombre_usuario(request.user), len(encabezados))
    ws.freeze_panes = f'A{fila_enc + 1}'

    for i, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=fila_enc, column=i, value=enc)
        cell.style = cabecera
    ws.row_dimensions[fila_enc].height = 30

    data_row = fila_enc + 1
    for c in clientes:
        nombre = c.razon_social if c.tipo_identificacion == 'NIT' else f'{c.nombres} {c.apellidos}'
        row_data = [
            f'{c.tipo_identificacion} {c.identificacion}',
            nombre,
            c.telefono,
            c.ciudad,
            c.estado,
            _nombre_usuario(c.creado_por),
            _formato_fecha(c.fecha_creacion),
            _nombre_usuario(c.modificado_por),
            _formato_fecha(c.modificado_en),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.border = borde
            cell.alignment = Alignment(vertical='center')
        data_row += 1

    _ajustar_columnas(ws)

    wb.properties.title   = 'Reporte de Clientes — SOLGASES'
    wb.properties.creator = _nombre_usuario(request.user)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    return response


@login_requerido
def exportar_proveedores_excel(request):
    """Genera y descarga el reporte Excel de proveedores con auditoría."""
    proveedores = Proveedor.objects.select_related('creado_por', 'modificado_por').order_by('identificacion')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proveedores'

    cabecera, dato, borde, _ = _estilo_excel(wb)

    encabezados = [
        'Identificación', 'Nombre / Razón social', 'Teléfono', 'Ciudad',
        'Estado', 'Creado por', 'Fecha creación', 'Modificado por', 'Última modificación'
    ]
    fila_enc = _insertar_encabezado(ws, 'REPORTE DE PROVEEDORES', _nombre_usuario(request.user), len(encabezados))
    ws.freeze_panes = f'A{fila_enc + 1}'

    for i, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=fila_enc, column=i, value=enc)
        cell.style = cabecera
    ws.row_dimensions[fila_enc].height = 30

    data_row = fila_enc + 1
    for p in proveedores:
        nombre = p.razon_social if p.tipo_identificacion == 'NIT' else f'{p.nombres} {p.apellidos}'
        row_data = [
            f'{p.tipo_identificacion} {p.identificacion}',
            nombre,
            p.telefono,
            p.ciudad,
            p.estado,
            _nombre_usuario(p.creado_por),
            _formato_fecha(p.fecha_creacion),
            _nombre_usuario(p.modificado_por),
            _formato_fecha(p.modificado_en),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.border = borde
            cell.alignment = Alignment(vertical='center')
        data_row += 1

    _ajustar_columnas(ws)

    wb.properties.title   = 'Reporte de Proveedores — SOLGASES'
    wb.properties.creator = _nombre_usuario(request.user)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'proveedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    return response