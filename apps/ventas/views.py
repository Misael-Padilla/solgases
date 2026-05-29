from io import BytesIO
from datetime import datetime

from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse
from django.views.decorators.http import require_POST

from apps.ventas.models import FacturaVenta
from apps.ventas.forms import FacturaVentaForm, DetalleVentaFormSet
from apps.productos.models import Producto, HistorialStock
from apps.insumos.models import Insumo, HistorialStockInsumo
from apps.usuarios.decoradores import login_requerido, admin_requerido
from apps.usuarios.models import HistorialCambio
from apps.usuarios.views import (
    _estilo_excel, _ajustar_columnas,
    _nombre_usuario, _formato_fecha, _insertar_encabezado,
)

_POR_PAGINA = 15


@login_requerido
def lista_ventas(request):
    """Lista todas las facturas de venta — accesible para ADMIN y EMP."""
    q = request.GET.get('q', '').strip()
    ventas = FacturaVenta.objects.all().order_by('-fecha_registro')
    if q:
        ventas = ventas.filter(
            Q(numero_factura__icontains=q) |
            Q(cliente__nombres__icontains=q) |
            Q(cliente__apellidos__icontains=q) |
            Q(cliente__razon_social__icontains=q)
        )
    paginator = Paginator(ventas, _POR_PAGINA)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Ventas', 'url': None},
    ]
    return render(request, 'ventas/lista_ventas.html', {
        'ventas':             page_obj,
        'page_obj':           page_obj,
        'page_range':         list(paginator.get_elided_page_range(page_obj.number, on_each_side=2, on_ends=1)),
        'paginator_ellipsis': paginator.ELLIPSIS,
        'q':                  q,
        'breadcrumbs':        breadcrumbs,
    })


@login_requerido
def detalle_venta(request, id):
    """Muestra el detalle de una factura de venta específica."""
    venta = get_object_or_404(FacturaVenta, id=id)
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Ventas', 'url': reverse('ventas:lista_ventas')},
        {'nombre': venta.numero_factura, 'url': None},
    ]
    return render(request, 'ventas/detalle_venta.html', {
        'venta': venta, 'breadcrumbs': breadcrumbs
    })


@login_requerido
def crear_venta(request):
    """Registra una nueva factura de venta con sus detalles."""
    if request.method == 'POST':
        form    = FacturaVentaForm(request.POST)
        formset = DetalleVentaFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    venta = form.save(commit=False)
                    venta.registrado_por = request.user
                    venta.save()
                    formset.instance = venta
                    formset.save()
                    messages.success(request, f'Factura {venta.numero_factura} registrada correctamente.')
                    return redirect('ventas:detalle_venta', id=venta.id)
            except ValueError as e:
                messages.error(request, str(e))
    else:
        form    = FacturaVentaForm()
        formset = DetalleVentaFormSet()
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Ventas', 'url': reverse('ventas:lista_ventas')},
        {'nombre': 'Registrar factura', 'url': None},
    ]
    return render(request, 'ventas/form_venta.html', {
        'form':    form,
        'formset': formset,
        'titulo':  'Registrar factura de venta',
        'breadcrumbs': breadcrumbs,
    })


@admin_requerido
@require_POST
def cambiar_estado_venta(request, id):
    """
    Activa o desactiva una factura — solo ADMIN. Requiere observación.
    Al desactivar: restaura el stock de todos los ítems (Pendiente Fase 5).
    Al activar: valida stock disponible y re-aplica el descuento.
    """
    venta = get_object_or_404(FacturaVenta, id=id)
    observacion = request.POST.get('observacion', '').strip()
    if not observacion:
        messages.error(request, 'La observación es obligatoria para cambiar el estado.')
        return redirect('ventas:lista_ventas')

    # Validar stock antes de reactivar (el descuento se re-aplica)
    if venta.estado == 'INACTIVO':
        for detalle in venta.detalles.all():
            if detalle.tipo_item == 'PRODUCTO':
                try:
                    p = Producto.objects.get(codigo=detalle.codigo_item)
                    if p.stock < detalle.cantidad:
                        messages.error(request, f'No se puede reactivar. Stock insuficiente para {p.nombre}: disponible {p.stock}, necesario {detalle.cantidad}.')
                        return redirect('ventas:lista_ventas')
                except Producto.DoesNotExist:
                    messages.error(request, f'Producto {detalle.codigo_item} no existe en el sistema.')
                    return redirect('ventas:lista_ventas')
            elif detalle.tipo_item == 'INSUMO':
                try:
                    i = Insumo.objects.get(codigo=detalle.codigo_item)
                    if i.stock < detalle.cantidad:
                        messages.error(request, f'No se puede reactivar. Stock insuficiente para {i.nombre}: disponible {i.stock}, necesario {detalle.cantidad}.')
                        return redirect('ventas:lista_ventas')
                except Insumo.DoesNotExist:
                    messages.error(request, f'Insumo {detalle.codigo_item} no existe en el sistema.')
                    return redirect('ventas:lista_ventas')

    with transaction.atomic():
        if venta.estado == 'ACTIVO':
            venta.estado = 'INACTIVO'
            accion    = 'DESACTIVAR'
            delta     = 1    # restaurar stock (sumar de vuelta)
            msg_stock = 'restaurado'
        else:
            venta.estado = 'ACTIVO'
            accion    = 'ACTIVAR'
            delta     = -1   # re-aplicar venta (restar de nuevo)
            msg_stock = 'revertido'

        tipo_mov = 'anulación' if delta == 1 else 'reactivación'
        for detalle in venta.detalles.all():
            motivo_stock = (
                f'{"Restauración" if delta == 1 else "Re-aplicación"} por '
                f'{tipo_mov} de venta {venta.numero_factura}: {observacion}'
            )
            if detalle.tipo_item == 'PRODUCTO':
                try:
                    producto  = Producto.objects.get(codigo=detalle.codigo_item)
                    stock_ant = producto.stock
                    producto.stock = max(0, producto.stock + delta * detalle.cantidad)
                    producto.save()
                    HistorialStock.objects.create(
                        producto=producto, tipo='MANUAL',
                        stock_anterior=stock_ant,
                        stock_nuevo=producto.stock,
                        motivo=motivo_stock,
                        realizado_por=request.user,
                    )
                except Producto.DoesNotExist:
                    pass
            elif detalle.tipo_item == 'INSUMO':
                try:
                    insumo    = Insumo.objects.get(codigo=detalle.codigo_item)
                    stock_ant = insumo.stock
                    insumo.stock = max(0, insumo.stock + delta * detalle.cantidad)
                    insumo.save()
                    HistorialStockInsumo.objects.create(
                        insumo=insumo, tipo='MANUAL',
                        stock_anterior=stock_ant,
                        stock_nuevo=insumo.stock,
                        motivo=motivo_stock,
                        realizado_por=request.user,
                    )
                except Insumo.DoesNotExist:
                    pass

        venta.save()
        HistorialCambio.objects.create(
            modelo='VENTA',
            objeto_id=venta.id,
            objeto_nombre=venta.numero_factura,
            accion=accion,
            observacion=observacion,
            realizado_por=request.user,
        )

    verb = 'desactivada' if accion == 'DESACTIVAR' else 'activada'
    messages.success(request, f'Factura {venta.numero_factura} {verb}. Stock {msg_stock}.')
    return redirect('ventas:lista_ventas')


@login_requerido
def exportar_ventas_excel(request):
    """Genera y descarga el reporte Excel de facturas de venta."""
    import openpyxl
    from openpyxl.styles import Alignment

    ventas = FacturaVenta.objects.select_related(
        'cliente', 'registrado_por'
    ).order_by('-fecha_registro')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    cabecera, dato, borde, _ = _estilo_excel(wb)

    encabezados = [
        'N° Factura', 'Cliente', 'Registrado por', 'Recibido por',
        'Fecha factura', 'Fecha registro', 'Método de pago',
        'Subtotal', 'IVA (%)', 'IVA ($)', 'Total',
        'Estado', 'Observaciones',
    ]
    fila_enc = _insertar_encabezado(
        ws, 'REPORTE DE VENTAS', _nombre_usuario(request.user), len(encabezados)
    )
    ws.freeze_panes = f'A{fila_enc + 1}'

    for i, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=fila_enc, column=i, value=enc)
        cell.style = cabecera
    ws.row_dimensions[fila_enc].height = 30

    data_row = fila_enc + 1
    for v in ventas:
        row_data = [
            v.numero_factura,
            str(v.cliente),
            _nombre_usuario(v.registrado_por),
            v.recibido_por or '—',
            _formato_fecha(v.fecha_factura),
            _formato_fecha(v.fecha_registro),
            v.get_metodo_pago_display(),
            float(v.subtotal),
            float(v.iva_porcentaje),
            float(v.iva),
            float(v.total),
            v.estado,
            v.observaciones or '—',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.border = borde
            cell.alignment = Alignment(vertical='center')
        data_row += 1

    _ajustar_columnas(ws)
    wb.properties.title   = 'Reporte de Ventas — SOLGASES'
    wb.properties.creator = _nombre_usuario(request.user)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    return response
