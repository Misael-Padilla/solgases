from io import BytesIO
from datetime import datetime

from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse
from django.views.decorators.http import require_POST

from apps.insumos.models import Insumo, HistorialStockInsumo
from apps.insumos.forms import InsumoForm, StockInsumoForm
from apps.usuarios.decoradores import login_requerido, admin_requerido
from apps.usuarios.models import HistorialCambio
from apps.usuarios.views import (
    _estilo_excel, _ajustar_columnas,
    _nombre_usuario, _formato_fecha, _insertar_encabezado,
)

_POR_PAGINA = 15


@login_requerido
def lista_insumos(request):
    """Lista todos los insumos — accesible para ADMIN y EMP."""
    q = request.GET.get('q', '').strip()
    insumos = Insumo.objects.all().order_by('codigo')
    if q:
        insumos = insumos.filter(
            Q(codigo__icontains=q) |
            Q(nombre__icontains=q) |
            Q(subcategoria__icontains=q) |
            Q(proveedor__razon_social__icontains=q) |
            Q(proveedor__nombres__icontains=q)
        )
    paginator = Paginator(insumos, _POR_PAGINA)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Insumos', 'url': None},
    ]
    return render(request, 'insumos/lista_insumos.html', {
        'insumos':            page_obj,
        'page_obj':           page_obj,
        'page_range':         list(paginator.get_elided_page_range(page_obj.number, on_each_side=2, on_ends=1)),
        'paginator_ellipsis': paginator.ELLIPSIS,
        'q':                  q,
        'breadcrumbs':        breadcrumbs,
    })


@login_requerido
def detalle_insumo(request, id):
    """Muestra el detalle de un insumo específico."""
    insumo    = get_object_or_404(Insumo, id=id)
    historial = HistorialStockInsumo.objects.filter(
        insumo=insumo
    ).select_related('realizado_por')[:10]
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Insumos', 'url': reverse('insumos:lista_insumos')},
        {'nombre': insumo.nombre, 'url': None},
    ]
    return render(request, 'insumos/detalle_insumo.html', {
        'insumo': insumo, 'historial': historial, 'breadcrumbs': breadcrumbs
    })


@login_requerido
def crear_insumo(request):
    """Crea un nuevo insumo — accesible para ADMIN y EMP."""
    if request.method == 'POST':
        form = InsumoForm(request.POST, request.FILES, es_edicion=False)
        if form.is_valid():
            insumo = form.save(commit=False)
            insumo.creado_por = request.user
            insumo.save()
            messages.success(request, 'Insumo creado correctamente.')
            return redirect('insumos:lista_insumos')
    else:
        form = InsumoForm(es_edicion=False)
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Insumos', 'url': reverse('insumos:lista_insumos')},
        {'nombre': 'Crear insumo', 'url': None},
    ]
    return render(request, 'insumos/form_insumo.html', {
        'form': form, 'titulo': 'Crear insumo', 'breadcrumbs': breadcrumbs
    })


@login_requerido
def editar_insumo(request, id):
    """Edita un insumo existente — accesible para ADMIN y EMP."""
    insumo = get_object_or_404(Insumo, id=id)
    if request.method == 'POST':
        form = InsumoForm(request.POST, request.FILES, instance=insumo, es_edicion=True)
        if form.is_valid():
            i = form.save(commit=False)
            i.modificado_por = request.user
            i.save()
            messages.success(request, 'Insumo actualizado correctamente.')
            return redirect('insumos:detalle_insumo', id=insumo.id)
    else:
        form = InsumoForm(instance=insumo, es_edicion=True)
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Insumos', 'url': reverse('insumos:lista_insumos')},
        {'nombre': insumo.nombre, 'url': reverse('insumos:detalle_insumo', args=[insumo.id])},
        {'nombre': 'Editar', 'url': None},
    ]
    return render(request, 'insumos/form_insumo.html', {
        'form': form, 'titulo': 'Editar insumo', 'breadcrumbs': breadcrumbs
    })


@admin_requerido
@require_POST
def cambiar_estado_insumo(request, id):
    """Activa o desactiva un insumo — solo ADMIN. Requiere observación."""
    insumo = get_object_or_404(Insumo, id=id)
    observacion = request.POST.get('observacion', '').strip()
    if not observacion:
        messages.error(request, 'La observación es obligatoria para cambiar el estado.')
        return redirect('insumos:lista_insumos')
    if insumo.estado == 'ACTIVO':
        insumo.estado = 'INACTIVO'
        accion = 'DESACTIVAR'
        messages.success(request, f'Insumo {insumo.nombre} desactivado.')
    else:
        insumo.estado = 'ACTIVO'
        accion = 'ACTIVAR'
        messages.success(request, f'Insumo {insumo.nombre} activado.')
    insumo.modificado_por = request.user
    insumo.save()
    HistorialCambio.objects.create(
        modelo='INSUMO',
        objeto_id=insumo.id,
        objeto_nombre=insumo.nombre,
        accion=accion,
        observacion=observacion,
        realizado_por=request.user,
    )
    return redirect('insumos:lista_insumos')


@login_requerido
def exportar_insumos_excel(request):
    """Genera y descarga el reporte Excel de insumos."""
    import openpyxl
    from openpyxl.styles import Alignment

    insumos = Insumo.objects.select_related(
        'proveedor', 'creado_por', 'modificado_por'
    ).order_by('codigo')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Insumos'

    cabecera, dato, borde, _ = _estilo_excel(wb)

    encabezados = [
        'Código', 'Nombre', 'Subcategoría', 'Unidad de medida',
        'Precio compra', 'Stock', 'Stock mínimo', 'Stock máximo', 'Nivel stock',
        'Proveedor', 'Estado', 'Fecha registro',
        'Creado por', 'Modificado por', 'Última modificación',
    ]
    fila_enc = _insertar_encabezado(
        ws, 'REPORTE DE INSUMOS', _nombre_usuario(request.user), len(encabezados)
    )
    ws.freeze_panes = f'A{fila_enc + 1}'

    for i, enc in enumerate(encabezados, 1):
        cell = ws.cell(row=fila_enc, column=i, value=enc)
        cell.style = cabecera
    ws.row_dimensions[fila_enc].height = 30

    data_row = fila_enc + 1
    for i in insumos:
        row_data = [
            i.codigo,
            i.nombre,
            i.get_subcategoria_display(),
            i.unidad_medida,
            float(i.precio_compra),
            i.stock,
            i.stock_minimo,
            i.stock_maximo,
            i.nivel_stock,
            str(i.proveedor),
            i.estado,
            _formato_fecha(i.fecha_creacion),
            _nombre_usuario(i.creado_por),
            _nombre_usuario(i.modificado_por),
            _formato_fecha(i.modificado_en),
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.border = borde
            c.alignment = Alignment(vertical='center')
        data_row += 1

    _ajustar_columnas(ws)
    wb.properties.title   = 'Reporte de Insumos — SOLGASES'
    wb.properties.creator = _nombre_usuario(request.user)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = f'insumos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, private'
    return response


@admin_requerido
def modificar_stock_insumo(request, id):
    """Modifica el stock manualmente — solo ADMIN. Requiere motivo obligatorio."""
    insumo = get_object_or_404(Insumo, id=id)
    if request.method == 'POST':
        form = StockInsumoForm(request.POST)
        if form.is_valid():
            stock_anterior        = insumo.stock
            insumo.stock          = form.cleaned_data['stock']
            insumo.modificado_por = request.user
            insumo.save()
            HistorialStockInsumo.objects.create(
                insumo=insumo,
                tipo='MANUAL',
                stock_anterior=stock_anterior,
                stock_nuevo=insumo.stock,
                motivo=form.cleaned_data['motivo'],
                realizado_por=request.user,
            )
            messages.success(request, f'Stock de {insumo.nombre} actualizado a {insumo.stock}.')
            return redirect('insumos:detalle_insumo', id=insumo.id)
    else:
        form = StockInsumoForm(initial={'stock': insumo.stock})
    breadcrumbs = [
        {'nombre': 'Dashboard', 'url': reverse('core:inicio')},
        {'nombre': 'Insumos', 'url': reverse('insumos:lista_insumos')},
        {'nombre': insumo.nombre, 'url': reverse('insumos:detalle_insumo', args=[insumo.id])},
        {'nombre': 'Modificar stock', 'url': None},
    ]
    return render(request, 'insumos/form_stock_insumo.html', {
        'form': form, 'insumo': insumo, 'breadcrumbs': breadcrumbs
    })
